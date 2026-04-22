#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build literature-register.xlsx with S.N., Author, Year, Title, Publication, Citations, Remarks."""
from __future__ import annotations

import json
import re
import time
import urllib.parse
import urllib.request
from html.parser import HTMLParser
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pypdf import PdfReader

ROOT = Path(__file__).resolve().parent
OPENALEX = "https://api.openalex.org"
USER_AGENT = "LiteratureRegister/1.0 (thesis; mailto:local)"

DOI_RE = re.compile(r"10\.\d{4,}/[^\s\]\)\"'<>]+", re.I)
YEAR_IN_NAME = re.compile(r"^([A-Za-z'`-]+)(\d{4})-(.+)\.(pdf|html)$")


def http_get_json(url: str, timeout: float = 45.0) -> dict | None:
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT, "Accept": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return json.loads(resp.read().decode("utf-8", errors="replace"))
    except Exception:
        return None


def extract_year_from_filename(name: str) -> int | None:
    m = YEAR_IN_NAME.match(name)
    if m:
        return int(m.group(2))
    return None


def pdf_text(path: Path, max_pages: int = 3) -> str:
    r = PdfReader(str(path))
    t = []
    for i in range(min(max_pages, len(r.pages))):
        t.append(r.pages[i].extract_text() or "")
    return "\n".join(t)


def first_doi(text: str) -> str | None:
    for m in DOI_RE.finditer(text):
        d = m.group(0).rstrip(".,;)")
        if len(d) > 8:
            return d
    return None


def meta_title_author(path: Path) -> tuple[str, str]:
    try:
        r = PdfReader(str(path))
        m = r.metadata or {}
        title = (m.get("/Title") or "").strip()
        author = (m.get("/Author") or "").strip()
        if title.startswith("Microsoft") or title.startswith("Springer") or len(title) < 6:
            title = ""
        return title, author
    except Exception:
        return "", ""


def format_authors_from_openalex(work: dict) -> str:
    authorships = work.get("authorships") or []
    names = []
    for a in authorships[:8]:
        author = (a.get("author") or {}).get("display_name")
        if author:
            names.append(author)
    if not names:
        return ""
    if len(names) <= 2:
        return " & ".join(names)
    return f"{names[0]} et al."


def openalex_by_doi(doi: str) -> dict | None:
    # doi may be 10.xxx/yyy
    doi_url = doi if doi.startswith("http") else f"https://doi.org/{doi}"
    wid = urllib.parse.quote(doi_url, safe="")
    return http_get_json(f"{OPENALEX}/works/{wid}")


def openalex_search(title: str, year: int | None) -> dict | None:
    q = re.sub(r"\s+", " ", title).strip()[:280]
    if not q:
        return None
    params = {"search": q, "per-page": 5}
    if year:
        params["filter"] = f"publication_year:{year}"
    url = f"{OPENALEX}/works?{urllib.parse.urlencode(params)}"
    data = http_get_json(url)
    if not data or not data.get("results"):
        return None
    return data["results"][0]


_BAD_LINE = re.compile(
    r"^(received|revised|accepted|published|issn|vol\.|issue|doi:|http|abstract|keywords|copyright|"
    r"article|page|\d+\s*$|introduction)",
    re.I,
)


def extract_title_by_patterns(text: str) -> str:
    patterns = [
        r"(Understanding the Key Factors Influencing Cybersecurity Practices in Nepalese Organizations)",
        r"(Assessing the Cybersecurity Literacy Proficiency among Bachelor(?:\u2019|')?s Degree Students in Nepal)",
        r"(Assessing the Cybersecurity Literacy Proficiency among[\s\S]{5,120}?Nepal)",
    ]
    for pat in patterns:
        m = re.search(pat, text, re.I)
        if m:
            return re.sub(r"\s+", " ", m.group(1)).strip()
    return ""


def title_for_search(meta_title: str, text: str) -> str:
    pat_title = extract_title_by_patterns(text)
    if pat_title:
        return pat_title
    if meta_title and len(meta_title) > 12 and not meta_title.lower().startswith("microsoft"):
        return re.sub(r"\s+", " ", meta_title).strip()

    # Title block often appears just before "DOI:" (incl. multi-line titles)
    low = text
    for marker in ("DOI:", "doi:", "Doi:"):
        if marker in low:
            pre = low.split(marker, 1)[0]
            chunk = pre[-1200:] if len(pre) > 1200 else pre
            lines = [ln.strip() for ln in chunk.splitlines() if ln.strip()]
            title_lines: list[str] = []
            for ln in reversed(lines):
                if len(ln) < 12:
                    continue
                if _BAD_LINE.search(ln):
                    break
                if "ISSN" in ln or "www." in ln.lower() or "http" in ln.lower():
                    continue
                if re.search(r"Journal of Management,\s*Technology\s*&\s*Social\s*Sciences", ln, re.I):
                    continue
                if re.fullmatch(r"[\d\s\-\u2013\u2014]+", ln):
                    continue
                title_lines.append(ln)
                if len(title_lines) >= 3:
                    break
            if title_lines:
                title_lines.reverse()
                cand = " ".join(title_lines)
                cand = re.sub(r"\s+", " ", cand).strip()
                if 25 <= len(cand) <= 400:
                    return cand

    # Longest plausible single line (avoid metadata / receipt lines)
    lines = []
    for ln in text.splitlines():
        s = ln.strip()
        if 25 <= len(s) <= 220 and not _BAD_LINE.search(s) and "http" not in s.lower():
            lines.append(s)
    if not lines:
        return ""
    lines.sort(key=len, reverse=True)
    return lines[0]


def guess_first_author_line(text: str) -> str:
    """Best-effort first author when PDF metadata is empty (first named author only)."""
    for line in text.splitlines():
        s = line.strip()
        if "," in s and re.search(r"\d+\*?\s*$", s.split(",")[0]):
            s = s.split(",")[0].strip()
        m = re.match(r"^([A-Z][a-zA-Z'.-]+(?:\s+[A-Z][a-zA-Z'.-]+){1,4})\d+[a-z]*\*?\s*$", s)
        if m:
            return m.group(1).strip()
    m = re.search(
        r"^([A-Z][a-zA-Z'.-]+(?:\s+[A-Z][a-zA-Z'.-]+){1,4})\d+[a-z]*\*",
        text,
        re.M,
    )
    if m:
        return m.group(1).strip()
    return ""


def title_word_overlap(a: str, b: str) -> float:
    sa = {w.lower() for w in re.findall(r"[A-Za-z]{3,}", a)}
    sb = {w.lower() for w in re.findall(r"[A-Za-z]{3,}", b)}
    if not sa or not sb:
        return 0.0
    return len(sa & sb) / len(sa | sb)


def guess_publication_from_text(text: str) -> str:
    head = text[:3500]
    m = re.search(
        r"(The\s+[^\n]+Journal[^\n]{5,120}|International\s+Journal\s+of\s+[^\n]{5,120}|"
        r"Journal\s+of\s+[A-Z][^\n]{5,100})",
        head,
        re.I,
    )
    if m:
        return re.sub(r"\s+", " ", m.group(1)).strip()
    return ""


def read_html_title(path: Path) -> str:
    class P(HTMLParser):
        def __init__(self) -> None:
            super().__init__()
            self.in_title = False
            self.buf: list[str] = []

        def handle_starttag(self, tag, attrs):
            if tag.lower() == "title":
                self.in_title = True

        def handle_endtag(self, tag):
            if tag.lower() == "title":
                self.in_title = False

        def handle_data(self, data):
            if self.in_title:
                self.buf.append(data)

    raw = path.read_text(encoding="utf-8", errors="replace")
    p = P()
    p.feed(raw)
    t = "".join(p.buf).strip()
    return t or "Cybersecurity questionnaire (HTML instrument)"


def row_for_pdf(path: Path, rel: str) -> dict:
    text = pdf_text(path)
    doi = first_doi(text)
    meta_t, meta_a = meta_title_author(path)
    year_fn = extract_year_from_filename(path.name)
    title = title_for_search(meta_t, text)

    work: dict | None = None
    source = ""
    if doi:
        work = openalex_by_doi(doi)
        if work:
            source = "doi"
    if work is None and title and year_fn:
        time.sleep(0.25)
        cand = openalex_search(title, year_fn)
        if cand:
            py = cand.get("publication_year")
            oa_t = cand.get("display_name") or cand.get("title") or ""
            if py and abs(py - year_fn) <= 1 and title_word_overlap(title, oa_t) >= 0.12:
                work = cand
                source = "search"

    remarks: list[str] = [rel.replace("\\", "/")]
    if "base/" in rel.replace("\\", "/") and "Ahamed2024" in path.name:
        remarks.append("Duplicate PDF copy (same article as root folder).")

    if work:
        oa_title = work.get("display_name") or work.get("title") or title
        oa_year = work.get("publication_year") or year_fn
        venue = (work.get("primary_location") or {}).get("source") or {}
        publication = venue.get("display_name") or ""
        cites = work.get("cited_by_count")
        if cites is None:
            cites = "n/a"
        else:
            cites = int(cites)
        authors_oa = format_authors_from_openalex(work)
        authors_pdf = guess_first_author_line(text)
        n_auth = len(work.get("authorships") or [])
        if authors_pdf:
            multi = n_auth > 1 or text[:4500].count("*") >= 2
            authors = f"{authors_pdf} et al." if multi else authors_pdf
        else:
            authors = authors_oa or meta_a
        if authors_pdf and authors_oa and authors_pdf.split()[0] != authors_oa.split()[0]:
            remarks.append("Lead author taken from PDF (differs from OpenAlex contributor list).")
        remarks.append("Citations from OpenAlex (as of export date); verify for your thesis.")
        return {
            "author": authors,
            "year": oa_year,
            "title": oa_title,
            "publication": publication,
            "citations": cites,
            "remarks": "; ".join(remarks),
        }

    # Fallback: filename + metadata only
    remarks.append("OpenAlex match not found; fill publication/citations manually if needed.")
    short_title = title or meta_t or path.stem.replace("-", " ")
    pub_guess = guess_publication_from_text(text)
    auth = meta_a or guess_first_author_line(text)
    if auth and text[:4500].count("*") >= 2 and " et al." not in auth:
        auth = f"{auth} et al."
    return {
        "author": auth,
        "year": year_fn or "",
        "title": short_title,
        "publication": pub_guess,
        "citations": "n/a",
        "remarks": "; ".join(remarks),
    }


def row_for_html(path: Path, rel: str) -> dict:
    return {
        "author": "",
        "year": "",
        "title": read_html_title(path),
        "publication": "Local research instrument (not a journal article)",
        "citations": "n/a",
        "remarks": f"{rel}; questionnaire / survey HTML for thesis data collection.",
    }


def main() -> None:
    files = sorted(
        [p for p in ROOT.rglob("*") if p.is_file() and p.suffix.lower() in {".pdf", ".html"} and ".venv" not in p.parts],
        key=lambda p: (len(p.relative_to(ROOT).parts), str(p).lower()),
    )

    headers = (
        "S.N.",
        "Author",
        "Year",
        "Title",
        "Publication",
        "Citations",
        "Main core theme",
        "Status",
        "Remarks",
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Literature"
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)

    for i, path in enumerate(files, start=1):
        rel = str(path.relative_to(ROOT))
        if path.suffix.lower() == ".pdf":
            time.sleep(0.12)
            data = row_for_pdf(path, rel)
        else:
            data = row_for_html(path, rel)
        cit = data["citations"]
        if cit is None:
            cit = "n/a"
        ws.append(
            [
                i,
                data["author"],
                data["year"],
                data["title"],
                data["publication"],
                cit,
                "",
                "Not read",
                data["remarks"],
            ]
        )

    # Widen columns slightly
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = min(48, 14 + (4 if col == 4 else 0))
    ws.column_dimensions["G"].width = 28
    ws.column_dimensions["H"].width = 14

    dv = DataValidation(
        type="list",
        formula1='"Not read,Read,In progress,Skimmed"',
        allow_blank=True,
    )
    dv.error = "Choose a status from the list."
    dv.errorTitle = "Invalid status"
    ws.add_data_validation(dv)
    dv.add(f"H2:H{ws.max_row}")

    out = ROOT / "literature-register.xlsx"
    wb.save(out)
    print(f"Wrote {len(files)} rows to {out}")


if __name__ == "__main__":
    main()
